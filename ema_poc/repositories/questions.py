"""Question Repository: versioned, queryable question store (FR-1).

Every mutation writes a new version row; "current" is the highest-version row
per question_id. History is never destroyed (FR-103). Functions accept an
injectable `now` ISO-8601 UTC timestamp for deterministic tests.
"""

from __future__ import annotations

import csv
import sqlite3
from datetime import datetime, timezone

from ema_poc.models import Question


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _iso(value: datetime | str | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return value.isoformat()


def _question_from_row(row: sqlite3.Row) -> Question:
    return Question(**dict(row))


def _insert_version(conn: sqlite3.Connection, q: Question) -> None:
    conn.execute(
        """
        INSERT INTO questions (
            question_id, version, question_text, persona, therapeutic_area,
            brand_focus, domain, active, approval_status, approver_name,
            created_at, updated_at, deleted_at, delete_reason
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            q.question_id,
            q.version,
            q.question_text,
            q.persona.value,
            q.therapeutic_area,
            q.brand_focus,
            q.domain.value,
            int(q.active),
            q.approval_status.value,
            q.approver_name,
            _iso(q.created_at),
            _iso(q.updated_at),
            _iso(q.deleted_at),
            q.delete_reason,
        ),
    )
    conn.commit()


def get_version(
    conn: sqlite3.Connection, question_id: str, version: int
) -> Question | None:
    row = conn.execute(
        "SELECT * FROM questions WHERE question_id = ? AND version = ?",
        (question_id, version),
    ).fetchone()
    return _question_from_row(row) if row else None


def get_current(conn: sqlite3.Connection, question_id: str) -> Question | None:
    row = conn.execute(
        "SELECT * FROM questions WHERE question_id = ? ORDER BY version DESC LIMIT 1",
        (question_id,),
    ).fetchone()
    return _question_from_row(row) if row else None


def add_question(
    conn: sqlite3.Connection,
    *,
    question_id: str,
    question_text: str,
    persona: str,
    domain: str,
    therapeutic_area: str | None = None,
    brand_focus: str | None = None,
    now: str | None = None,
) -> Question:
    if get_current(conn, question_id) is not None:
        raise ValueError(f"Question already exists: {question_id}")
    now = now or _now_iso()
    q = Question(
        question_id=question_id,
        version=1,
        question_text=question_text,
        persona=persona,
        domain=domain,
        therapeutic_area=therapeutic_area,
        brand_focus=brand_focus,
        created_at=now,
        updated_at=now,
    )
    _insert_version(conn, q)
    return get_version(conn, question_id, 1)


def _enum_value(x):
    return x.value if hasattr(x, "value") else x


def list_questions(
    conn: sqlite3.Connection,
    *,
    persona=None,
    therapeutic_area: str | None = None,
    brand_focus: str | None = None,
    domain=None,
    active: bool | None = None,
    approval_status=None,
    include_deleted: bool = False,
) -> list[Question]:
    """Return the current version of each question, filtered. Excludes
    soft-deleted questions unless include_deleted=True."""
    sql = [
        "SELECT q.* FROM questions q",
        "JOIN (SELECT question_id, MAX(version) AS v FROM questions"
        " GROUP BY question_id) m",
        "ON q.question_id = m.question_id AND q.version = m.v",
    ]
    where: list[str] = []
    params: list = []
    if persona is not None:
        where.append("q.persona = ?")
        params.append(_enum_value(persona))
    if therapeutic_area is not None:
        where.append("q.therapeutic_area = ?")
        params.append(therapeutic_area)
    if brand_focus is not None:
        where.append("q.brand_focus = ?")
        params.append(brand_focus)
    if domain is not None:
        where.append("q.domain = ?")
        params.append(_enum_value(domain))
    if active is not None:
        where.append("q.active = ?")
        params.append(int(active))
    if approval_status is not None:
        where.append("q.approval_status = ?")
        params.append(_enum_value(approval_status))
    if not include_deleted:
        where.append("q.deleted_at IS NULL")
    if where:
        sql.append("WHERE " + " AND ".join(where))
    sql.append("ORDER BY q.question_id")
    rows = conn.execute("\n".join(sql), params).fetchall()
    return [_question_from_row(r) for r in rows]


def update_question(
    conn: sqlite3.Connection, question_id: str, *, now: str | None = None, **changes
) -> Question:
    """Write a new version with `changes` applied. `created_at` is preserved
    from the current version; `updated_at` is set to `now`. Raises KeyError if
    the question does not exist."""
    current = get_current(conn, question_id)
    if current is None:
        raise KeyError(f"No such question: {question_id}")
    data = current.model_dump()
    data.update(changes)
    data["version"] = current.version + 1
    data["updated_at"] = now or _now_iso()
    new = Question(**data)  # re-validates the applied changes
    _insert_version(conn, new)
    return get_version(conn, question_id, new.version)


def deactivate_question(
    conn: sqlite3.Connection, question_id: str, *, now: str | None = None
) -> Question:
    return update_question(conn, question_id, active=False, now=now)


def approve_question(
    conn: sqlite3.Connection,
    question_id: str,
    approver_name: str,
    *,
    now: str | None = None,
) -> Question:
    return update_question(
        conn,
        question_id,
        approval_status="APPROVED",
        approver_name=approver_name,
        now=now,
    )


def reject_question(
    conn: sqlite3.Connection,
    question_id: str,
    approver_name: str,
    *,
    now: str | None = None,
) -> Question:
    return update_question(
        conn,
        question_id,
        approval_status="REJECTED",
        approver_name=approver_name,
        now=now,
    )


def soft_delete_question(
    conn: sqlite3.Connection, question_id: str, reason: str, *, now: str | None = None
) -> Question:
    """Tombstone the question via a new version with deleted_at set and
    active=False. History is preserved (no physical delete; DM-003)."""
    now = now or _now_iso()
    return update_question(
        conn,
        question_id,
        now=now,
        deleted_at=now,
        delete_reason=reason,
        active=False,
    )


def history(conn: sqlite3.Connection, question_id: str) -> list[Question]:
    rows = conn.execute(
        "SELECT * FROM questions WHERE question_id = ? ORDER BY version ASC",
        (question_id,),
    ).fetchall()
    return [_question_from_row(r) for r in rows]


def active_approved(conn: sqlite3.Connection) -> list[Question]:
    """Current questions that are active AND approved AND not soft-deleted —
    the set the runner dispatches (SE-002, BR-009)."""
    return list_questions(conn, active=True, approval_status="APPROVED")


def _coerce_optional(value) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _upsert_row(conn: sqlite3.Connection, row: dict, now: str) -> None:
    qid = str(row["question_id"]).strip()
    fields = dict(
        question_text=str(row["question_text"]).strip(),
        persona=str(row["persona"]).strip(),
        domain=str(row["domain"]).strip(),
        therapeutic_area=_coerce_optional(row.get("therapeutic_area")),
        brand_focus=_coerce_optional(row.get("brand_focus")),
    )
    if get_current(conn, qid) is None:
        add_question(conn, question_id=qid, now=now, **fields)
    else:
        update_question(conn, qid, now=now, **fields)


def import_questions_csv(
    conn: sqlite3.Connection, path: str, *, now: str | None = None
) -> int:
    """Upsert questions from a CSV with columns: question_id, question_text,
    persona, domain, therapeutic_area, brand_focus. Existing question_ids are
    updated as a new version; new ones are added (FR-105). Returns row count."""
    now = now or _now_iso()
    with open(path, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    for row in rows:
        _upsert_row(conn, row, now)
    return len(rows)


def import_questions_excel(
    conn: sqlite3.Connection, path: str, *, now: str | None = None
) -> int:
    """Upsert questions from an .xlsx whose first row is the header (same
    columns as CSV import). Returns the number of data rows processed."""
    from openpyxl import load_workbook

    now = now or _now_iso()
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter))
        count = 0
        for values in rows_iter:
            if all(v is None for v in values):
                continue
            _upsert_row(conn, dict(zip(headers, values)), now)
            count += 1
        return count
    finally:
        wb.close()
